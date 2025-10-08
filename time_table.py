import pandas as pd
import tkinter as tk
from tkinter import filedialog
import os
import pickle
from collections import defaultdict
from datetime import datetime, timedelta
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def parse_time(t):
    return datetime.strptime(t, "%H:%M")

def format_time(dt):
    return dt.strftime("%H:%M")

def time_diff_str(t1, t2):
    """Return HH:MM string for difference between t1 and t2, handles overnight spans."""
    if t2 < t1:
        t2 += timedelta(days=1)
    delta = t2 - t1
    total_minutes = int(delta.total_seconds() // 60)
    hours = total_minutes // 60
    minutes = total_minutes % 60
    return f"{hours:02}:{minutes:02}"



def group_by_duty_name(tuple_list):
    grouped = defaultdict(list)
    for t in tuple_list:
        grouped[t[6]].append(t)
    return list(grouped.values())

def ensure_row_exists(df, row_idx):
    """Adds empty rows to df if row_idx is out of bounds."""
    while len(df) <= row_idx:
        df.loc[len(df)] = [None] * df.shape[1]


# -------------------- File Picker --------------------
root = tk.Tk()
root.withdraw()
file_path = filedialog.askopenfilename(title="Select Final Excel File", filetypes=[("Excel files", "*.xlsx *.xls")])
if not file_path:
    print("No file selected. Exiting.")
    exit()

# -------------------- Load and Fix Excel --------------------
df = pd.read_excel(file_path)

# Fill down merged cells (especially Duty Name)
df["Duty Name"] = df["Duty Name"].ffill()

required_cols = [
    "Origin", "Destination", "Start Time", "End Time", "Trip No",
    "Depot", "Duty Name", "Route Number"
]

missing = [col for col in required_cols if col not in df.columns]
if missing:
    print(f"Missing columns in the file: {missing}")
    exit()

# -------------------- Extract Tuples --------------------
tuples = []

for _, row in df.iterrows():
    origin = str(row["Origin"]).strip()
    dest = str(row["Destination"]).strip()
    dep_time = str(row["Start Time"]).strip().split(":")[0:2]
    arr_time = str(row["End Time"]).strip().split(":")[0:2]
    dep_time = ":".join(dep_time)
    arr_time = ":".join(arr_time)
    trip_no = int(row["Trip No"])
    depot = str(row["Depot"]).strip()
    duty_full = str(row["Duty Name"]).strip()
    route = str(row["Route Number"]).strip()

    # Remove route from Duty Name (e.g., 542/7A → 7A)
    if "/" in duty_full:
        _, duty = duty_full.split("/", 1)
    else:
        duty = duty_full

    tuples.append((origin, dest, dep_time, arr_time, trip_no, depot, duty, route))

# -------------------- Extract Stop Names --------------------

stops_set = {t[0] for t in tuples}
t1 = tuples[0]
depot = t1[5]

# Assume: origin_set is a set of strings, depot is a string
stops_set = {origin for origin in stops_set if depot not in origin}

# -------------------- Grouping Tuples --------------------

# Split tuples by presence of 'A' in duty_name
tuples_with_A = [t for t in tuples if 'A' in str(t[6])]
tuples_without_A = [t for t in tuples if 'A' not in str(t[6])]

# Apply grouping
grouped_without_A = group_by_duty_name(tuples_without_A)
grouped_with_A = group_by_duty_name(tuples_with_A)

# Final nested structure
nested_grouped_tuples = [grouped_without_A, grouped_with_A]

# -------------------- Creating Dataframe --------------------

# Step 1: Compute number of dynamic columns
static_cols = ["Duty Number", "Duty Hours", "Crew Sign In/Out Time", "Out/in Shedding"]
dynamic_cols = []

for origin in sorted(stops_set):
    dynamic_cols.append("Arrival")
    dynamic_cols.append("Departure")

# Step 2: Create full column list
all_columns = static_cols + dynamic_cols

# Step 3: Create empty DataFrame
df_custom = pd.DataFrame(columns=all_columns)

# -------------------- Filling Times in DF --------------------

all_duties = []

for duty_type in nested_grouped_tuples:  # [duties_without_A, duties_with_A]
    duties = []

    for duty_name_group in duty_type:  # Each duty_name_group is a list of tuples
        df = df_custom.copy()
        duty_name = duty_name_group[0][6]  # All tuples in group share same duty name

        col_idx = 4  # Start from first dynamic column
        row_idx = 0

        for i, tup in enumerate(duty_name_group):
            dep_time = tup[2]
            arr_time = tup[3]

            # First tuple: set sign-in and first arrival
            if i == 0:
                ensure_row_exists(df, row_idx)
                df.at[row_idx, "Out/in Shedding"] = dep_time
                df.iat[row_idx, col_idx] = arr_time
                col_idx += 1

            # Last tuple: next row, departure and sign-out
            elif i == len(duty_name_group) - 1:
                ensure_row_exists(df, row_idx)
                df.iat[row_idx, col_idx] = dep_time
                col_idx += 1
                if col_idx >= df.shape[1]:
                    row_idx += 1
                    col_idx = 4
                    ensure_row_exists(df, row_idx)
                df.at[row_idx, "Out/in Shedding"] = arr_time


            # Middle tuples: arrival/departure in sequence
            else:
                ensure_row_exists(df, row_idx)
                df.iat[row_idx, col_idx] = dep_time
                col_idx += 1

                if col_idx >= df.shape[1]:
                    row_idx += 1
                    col_idx = 4
                    ensure_row_exists(df, row_idx)

                df.iat[row_idx, col_idx] = arr_time
                col_idx += 1

                if col_idx >= df.shape[1]:
                    row_idx += 1
                    col_idx = 4
                    ensure_row_exists(df, row_idx)

        # Post-process each duty DataFrame
        first_row_idx = df.first_valid_index()
        last_row_idx = df.last_valid_index()

        if first_row_idx is not None and last_row_idx is not None:
            # 1. Insert Duty Name (tup[6]) in col 0 of first row
            df.iat[first_row_idx, 0] = duty_name_group[0][6]

            # 2. Set Crew Sign In time in first row (10 mins before col[3])
            raw_time = df.at[first_row_idx, "Out/in Shedding"]
            if pd.notna(raw_time):
                t = parse_time(raw_time)
                df.iat[first_row_idx, 2] = format_time(t - timedelta(minutes=10))

            # 3. Set Crew Sign Out time in last row (10 mins after col[3])
            raw_time = df.at[last_row_idx, "Out/in Shedding"]
            if pd.notna(raw_time):
                t = parse_time(raw_time)
                df.iat[last_row_idx, 2] = format_time(t + timedelta(minutes=10))

            # 4. Calculate Duty Hours = diff between first and last Out/in Shedding
            t_start = parse_time(df.at[first_row_idx, "Out/in Shedding"])
            t_end = parse_time(df.at[last_row_idx, "Out/in Shedding"])
            df.iat[last_row_idx, 1] = time_diff_str(t_start, t_end)

        duties.append(df)

    all_duties.append(duties)

# -------------------- Merging Dataframes --------------------

# Step 1: Merge DataFrames in each duty type group
merged_blocks = []
for duty_list in all_duties:
    block_df = pd.concat(duty_list, ignore_index=True)
    merged_blocks.append(block_df)

# Step 2: Create the separating row
separator = pd.DataFrame([[None] * merged_blocks[0].shape[1]], columns=merged_blocks[0].columns)
separator.iloc[0, 0] = "Evening Shifts"  # Set first column

# Step 3: Final concatenation
df_final_schedule = pd.concat([merged_blocks[0], separator, merged_blocks[1]], ignore_index=True)

# -------------------- Adding New Headings --------------------

# Step 1: Save current column names as first row of data
df_final_schedule.loc[-1] = df_final_schedule.columns  # insert old headers as first row
df_final_schedule.index = df_final_schedule.index + 1  # shift index
df_final_schedule = df_final_schedule.sort_index()     # reorder rows

# Step 2: Build new column names
new_columns = []
arrival_counter = 0
sorted_stops = sorted(stops_set)

for col in df_final_schedule.columns:
    if col == df_final_schedule.columns[0]:  # first column
        new_columns.append(depot)
    elif "Arrival" in col:
        new_columns.append(sorted_stops[arrival_counter])
        arrival_counter += 1
    else:
        new_columns.append("")

# Step 3: Replace the column headers
df_final_schedule.columns = new_columns


import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# --- 1. Define multi-level header ---
# Build first and second rows
first_row = []
second_row = []

static_cols = ["Depot", "", "", ""]  # e.g. "Duty Number", "Duty Hours", etc.
first_row.extend(static_cols)
second_row.extend(["Duty Number", "Duty Hours", "Crew Sign In/Out Time", "Out/in Shedding"])

for stop in sorted(stops_set):
    first_row.extend([stop, stop])
    second_row.extend(["Arrival", "Departure"])

# Convert DataFrame to use second_row as columns
df_final_schedule.columns = second_row
df_final_schedule.loc[-1] = first_row  # insert first header row as data row
df_final_schedule.index = df_final_schedule.index + 1
df_final_schedule = df_final_schedule.sort_index()

# --- 2. Write to Excel ---
output_excel = os.path.splitext(file_path)[0] + "_schedule.xlsx"
with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
    df_final_schedule.to_excel(writer, index=False, header=False, startrow=1)
    ws = writer.sheets['Sheet1']

    # --- 3. Merge header cells ---
    col = 1
    for group, count in zip(first_row, [1 if val == "" else 2 for val in second_row]):
        if group != "":
            merge_end = col + count - 1
            if merge_end > col:
                ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=merge_end)
                ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")
        col += 1

print(f"\n Final Excel saved to: {output_excel}")

