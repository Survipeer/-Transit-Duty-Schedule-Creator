import os
import re
import sys
import pandas as pd
import openpyxl
import tkinter as tk
from tkinter import filedialog
from datetime import datetime, timedelta
from tqdm import tqdm

# -------------------- Utility Functions --------------------

def is_hhmm(val):
    return isinstance(val, str) and re.fullmatch(r"\d{1,2}:\d{2}", val.strip())

def is_empty(val):
    return pd.isna(val) or str(val).strip() == ""

def get_stop_name(stop_row, col_idx):
    while col_idx >= 0:
        val = stop_row.iloc[col_idx]
        if pd.notna(val) and str(val).strip().lower() != 'nan':
            return str(val).strip()
        col_idx -= 1
    return "UNKNOWN_STOP"

def map_shift(duty_val):
    if pd.isna(duty_val):
        return ""
    duty_str = str(duty_val).strip()
    if re.fullmatch(r"\d+", duty_str):
        return "Day out 1"
    elif re.fullmatch(r"\d+[A]", duty_str):
        return "Day out 2"
    else:
        return ""

def map_bus_id(duty_name_val):
    if pd.isna(duty_name_val):
        return ""
    duty_str = str(duty_name_val).strip()
    if re.fullmatch(r"\d+", duty_str):
        return duty_str
    elif re.fullmatch(r"\d+[A]", duty_str):
        return re.match(r"(\d+)[A]", duty_str).group(1)
    else:
        return ""

def compute_run_time(start, end):
    try:
        fmt = "%H:%M"
        t1 = datetime.strptime(start, fmt)
        t2 = datetime.strptime(end, fmt)
        if t2 < t1:
            t2 += timedelta(days=1)
        return int((t2 - t1).total_seconds() // 60)
    except:
        return None

# -------------------- File Picker --------------------

root = tk.Tk()
root.withdraw()
file_path = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel files", "*.xlsx *.xls")])
if not file_path:
    print("No file selected. Exiting.")
    sys.exit(1)

excel_name = os.path.splitext(os.path.basename(file_path))[0]

# -------------------- Step 1: Extract Tables --------------------

print(f"\n Reading file: {file_path}")
wb = openpyxl.load_workbook(file_path, data_only=True)
dataframes = []
df_counter = 1

for sheetname in tqdm(wb.sheetnames, desc="Scanning sheets"):
    ws = wb[sheetname]
    rows = list(ws.iter_rows())
    i = 0

    while i < len(rows):
        found_header = False
        while i < len(rows):
            row_values = [str(cell.value).strip() if cell.value is not None else "" for cell in rows[i]]
            if "Duty Number" in row_values:
                header_row_idx = i
                duty_idx = row_values.index("Duty Number")
                arrival_count = sum("Arrival" in val for val in row_values)
                total_columns = 4 + 2 * arrival_count
                found_header = True
                break
            i += 1

        if not found_header:
            break

        start_row = max(0, header_row_idx - 2)
        i = header_row_idx + 1
        table_data = []

        for k in range(start_row, header_row_idx + 1):
            row = rows[k][duty_idx: duty_idx + total_columns]
            clean_row = [cell.value.strftime("%H:%M") if hasattr(cell.value, "strftime") else cell.value for cell in row]
            table_data.append(clean_row)

        empty_rows = 0
        while i < len(rows):
            row = rows[i][duty_idx: duty_idx + total_columns]
            clean_row = [cell.value.strftime("%H:%M") if hasattr(cell.value, "strftime") else cell.value for cell in row]
            if all(val is None or str(val).strip() == "" for val in clean_row):
                empty_rows += 1
                if empty_rows > 3:
                    break
            else:
                empty_rows = 0
                table_data.append(clean_row)
            i += 1

        if table_data:
            col_headers = [f"Col_{j+1}" for j in range(total_columns)]
            df = pd.DataFrame(table_data, columns=col_headers)
            dataframes.append(df)
            df_counter += 1

if not dataframes:
    print(" No valid tables found. Exiting.")
    sys.exit(1)

# -------------------- Step 2: Convert to Trip Tuples --------------------

print(f"\n Extracting trip tuples from {len(dataframes)} tables...")
all_tuples = []
for df in tqdm(dataframes, desc="Building tuples"):
    if df.shape[0] < 4:
        continue

    stop_name_row = df.iloc[1]
    head_row = df.iloc[2]
    depot = str(df.iloc[1].get("Col_1", "")).strip()
    route = str(df.iloc[0].get("Col_1", "")).strip()
    num_cols = df.shape[1]
    evening_seen = False

    i = 0
    while i < len(df):
        row = df.iloc[i]
        if row.astype(str).str.contains("Evening Duties", case=False, na=False).any():
            evening_seen = True

        col_1 = str(row.get("Col_1", "")).strip()
        block_start = None
        if re.fullmatch(r"\d+[A]?", col_1):
            block_start = i
        if block_start is None:
            i += 1
            continue

        block_end = block_start
        for j in range(block_start + 1, len(df)):
            check_row = df.iloc[j]
            if (
                all(is_hhmm(check_row.get(f"Col_{c}", "")) for c in range(2, 5)) and
                all(is_empty(check_row.get(f"Col_{c}", "")) for c in range(5, num_cols + 1))
            ):
                block_end = j
                break
        else:
            block_end = len(df) - 1

        duty_name = str(df.iloc[block_start].get("Col_1", "")).strip()
        if evening_seen and duty_name.isdigit():
            duty_name += 'A'

        cell_tuples = []
        for r in range(block_start, block_end + 1):
            row = df.iloc[r]
            for c in range(3, num_cols):
                col_key = f"Col_{c + 1}"
                time_val = row.get(col_key, "")
                if is_hhmm(time_val):
                    stop = get_stop_name(stop_name_row, c)
                    arr_dep = "a" if "Arrival" in str(head_row.get(col_key, "")).strip() else "d"
                    cell_tuples.append((time_val.strip(), stop, arr_dep))

        if cell_tuples:
            last_time, last_stop, last_arr_dep = cell_tuples[-1]
            if last_arr_dep == 'd':
                cell_tuples[-1] = (last_time, last_stop, 'a')

        trip_num = 1
        for j in range(0, len(cell_tuples) - 1, 2):
            first = cell_tuples[j]
            second = cell_tuples[j + 1]
            if first[2] == 'd' and second[2] == 'a':
                dep_time, start_stop = first[0], first[1]
                arr_time, end_stop = second[0], second[1]
            elif first[2] == 'a' and second[2] == 'd':
                dep_time, start_stop = second[0], second[1]
                arr_time, end_stop = first[0], first[1]
            else:
                continue

            if start_stop == "UNKNOWN_STOP" or end_stop == "UNKNOWN_STOP":
                continue

            all_tuples.append((start_stop, end_stop, dep_time, arr_time, trip_num, depot, duty_name, route))
            trip_num += 1
        i = block_end + 1

if not all_tuples:
    print(" No trip tuples could be formed. Exiting.")
    sys.exit(1)

# -------------------- Step 3: Sch kms Input --------------------

print("\n Enter Scheduled Kilometers (Sch kms) for each Origin → Destination pair:")
unique_od_pairs = sorted({(start, end) for start, end, *_ in all_tuples})
sch_kms_dict = {}

for origin, dest in unique_od_pairs:
    while True:
        try:
            val = input(f"  {origin} → {dest}: ").strip()
            sch_kms_dict[(origin, dest)] = float(val)
            break
        except ValueError:
            print("     Invalid input. Please enter a numeric value.")

# -------------------- Step 4: Build Final Output --------------------

required_cols = [
    "S.No", "Depot", "Trip No", "Duty Name", "Duty Working Day Type",
    "Route Number", "Route Direction", "Origin", "Destination",
    "Start Time", "End Time", "Trip Type", "Sch kms",
    "Run Time", "Shift", "Bus Id"
]

rows = []
for i, (start_stop, end_stop, dep_time, arr_time, trip_num, depot, duty_name, route) in enumerate(tqdm(all_tuples, desc="Creating final rows"), start=1):
    run_time = compute_run_time(dep_time, arr_time)
    row = {
        "S.No": i,
        "Depot": depot,
        "Trip No": trip_num,
        "Duty Name": f"{route}/{duty_name}",
        "Duty Working Day Type": "Monday to Sunday",
        "Route Number": route,
        "Route Direction": "",  # Can be customized
        "Origin": start_stop,
        "Destination": end_stop,
        "Start Time": f"{dep_time}:00",
        "End Time": f"{arr_time}:00",
        "Trip Type": "Regular Trip",
        "Sch kms": sch_kms_dict.get((start_stop, end_stop), ""),
        "Run Time": run_time,
        "Shift": map_shift(duty_name),
        "Bus Id": map_bus_id(duty_name)
    }
    rows.append(row)

df_final = pd.DataFrame(rows, columns=required_cols)

# -------------------- Step 5: Save Excel --------------------

from openpyxl.utils import get_column_letter

output_file = os.path.splitext(file_path)[0] + "_final_schedule.xlsx"
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    df_final.to_excel(writer, index=False, sheet_name="Sheet1")

    ws = writer.sheets["Sheet1"]
    ws.freeze_panes = ws["A2"]

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    # Merge Duty Name cells if consecutive rows have same value
    duty_col_idx = df_final.columns.get_loc("Duty Name") + 1  # 1-based index for Excel
    prev_value = None
    merge_start = 2  # Start from row 2 (row 1 is header)

    for i in range(2, len(df_final) + 2):  # Excel row numbers
        curr_value = ws.cell(i, duty_col_idx).value
        if curr_value != prev_value:
            if i - merge_start > 1:
                ws.merge_cells(start_row=merge_start, start_column=duty_col_idx,
                               end_row=i - 1, end_column=duty_col_idx)
            merge_start = i
        prev_value = curr_value

    # Check and merge the last group if needed
    if len(df_final) + 2 - merge_start > 1:
        ws.merge_cells(start_row=merge_start, start_column=duty_col_idx,
                       end_row=len(df_final) + 1, end_column=duty_col_idx)

print(f"\nExcel file saved at: {output_file}")
